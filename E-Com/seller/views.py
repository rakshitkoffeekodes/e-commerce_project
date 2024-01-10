import os
import datetime
from datetime import timedelta, time, date
import random
import openpyxl
import xlsxwriter
from django.contrib.auth import authenticate
from django.contrib.auth.hashers import check_password, make_password
from django.core.files.storage import default_storage
from django.db.models import Q
from django.http import JsonResponse
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.tokens import RefreshToken
from .serilizers import *
from django.conf import settings
from django.core.mail import send_mail
from base.models import *


# ==========This function is for generated OTP==========

def generate_otp(first_name, last_name, email):
    otp = random.randint(000000, 999999)
    subject = 'One Time Password (OTP) Confirmation'
    message = f'''Hi {first_name} {last_name},
                            This email confirms your one-time password (OTP) for E-Commerce Site.
                            Your OTP is: {otp}
                            Thank you,
                            E-Commerce Site'''
    email_from = settings.EMAIL_HOST_USER
    recipient_list = [email, ]
    send_mail(subject, message, email_from, recipient_list)
    return otp


# ==========This function is for registers==========

@api_view(['POST'])
def register(request):
    first_name = request.POST['first_name']
    last_name = request.POST['last_name']
    email = request.POST['email']
    mobile_no = request.POST['mobile_no']
    password = request.POST['password']
    confirm_password = request.POST['confirm_password']
    try:
        mobile_data = Register.objects.filter(mobile_no=mobile_no)
        email_data = User.objects.filter(email=email)
        if not email.endswith('.com'):
            return JsonResponse({'message': 'Invalid email address.'}, status=400)
        if len(mobile_no) != 10:
            return JsonResponse({'message': 'Invalid mobile number.'}, status=400)
        if len(mobile_data) > 0 or len(email_data) > 0:
            return JsonResponse({'message': 'User already exists.'}, status=400)
        if password != confirm_password:
            return JsonResponse({'message': 'Passwords do not match.'}, status=400)
        global user_otp, temp
        user_otp = generate_otp(first_name, last_name, email)
        temp = {
            "first_name": first_name,
            "last_name": last_name,
            "email": email,
            "mobile_no": mobile_no,
            "password": password,
        }
        return JsonResponse(
            {'message': 'OTP sent successfully. Please check your email for the OTP.'}, status=200)
    except Exception as e:
        return JsonResponse({'message': e.__str__()}, status=400)


# ==========This function is for otp verification==========

@api_view(['POST'])
def otp_verification(request):
    otp = request.POST['otp']
    try:
        if user_otp == int(otp):
            user = User.objects.create(
                username=temp["email"].split('@')[0],
                email=temp["email"],
                first_name=temp["first_name"],
                last_name=temp["last_name"],
                password=make_password(temp["password"])
            )
            user_register = Register(
                register_user=user,
                mobile_no=temp["mobile_no"],
            )
            user_register.save()
            return JsonResponse({'message': 'Register Success'}, status=200)
        else:
            return JsonResponse({'message': 'OTP is not match.'}, status=400)
    except Exception as e:
        return JsonResponse({'message': str(e)}, status=400)


# ==========This function is for login==========

@api_view(['POST'])
def login(request):
    email = request.POST['email']
    password = request.POST['password']

    try:
        user = User.objects.get(email=email)
        if check_password(password, user.password):
            auth = authenticate(username=user.username, password=password)
            if auth is not None:
                refresh = RefreshToken.for_user(user)
                return JsonResponse(
                    {'message': 'You are successfully logged in', 'refresh': str(refresh),
                     'access': str(refresh.access_token)},
                    status=200)
            else:
                return JsonResponse({'message': 'User credentials were not provided.'}, status=400)
        else:
            return JsonResponse({'message': 'Please enter the correct password'}, status=400)
    except Register.DoesNotExist:
        return JsonResponse({'message': 'Please enter the correct email'}, status=400)
    except Exception as e:
        return JsonResponse({'message': e.__str__()}, status=400)


# ==========This function is for profile==========

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def profile(request):
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            serializer = RegisterSerializer(seller_user)
            serializer_data = serializer.data.get('register_user')
            serializer_dictionary = serializer_data
            for data in serializer.data:
                if 'register_user' not in data:
                    serializer_dictionary[data] = serializer.data.get(data)
            return JsonResponse({'message': 'Profile Data', 'profile_data': serializer_dictionary}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'profile_data': None}, status=400)


# ==========This function is for update profile==========

@api_view(['PUT'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def update_profile(request):
    profile_picture = request.FILES.get('profile_image', '')
    first_name = request.POST.get('first_name', '')
    last_name = request.POST.get('last_name', '')
    mobile_no = request.POST.get('mobile_no', '')
    address = request.POST.get('address', '')
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            if not profile_picture == '':
                seller_user.profile_picture = profile_picture
            if not first_name == '':
                user.first_name = first_name
            if not last_name == '':
                user.last_name = last_name
            if not mobile_no == '':
                seller_user.mobile_no = mobile_no
            if not address == '':
                seller_user.address = address
            user.save()
            seller_user.save()
            return JsonResponse({'message': 'Profile Update Successfully.'}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__()}, status=400)


# ==========This function is to insert the product==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def add_product(request):
    product_category = request.POST['product_category'].upper()
    product_sub_category = request.POST['product_sub_category'].upper()
    product_images = request.FILES.getlist('product_images')
    product_SKU = request.POST['product_SKU']
    product_name = request.POST['product_name']
    product_price = request.POST['product_price']
    product_sale_price = request.POST['product_sale_price']
    product_quantity = request.POST['product_quantity']
    product_branding = request.POST['product_branding']
    product_tags = request.POST['product_tags']
    product_size = request.POST['product_size']
    product_color = request.POST['product_color']
    product_fabric = request.POST['product_fabric']
    product_description = request.POST['product_description']
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            product_id = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "A", "B", "C", "D", "E", "F", "G", "H"]
            products_id = random.choices(product_id, k=9)
            product = "".join(products_id)
            product_image = [
                default_storage.save(os.path.join(settings.MEDIA_ROOT, 'Product', image.name.replace(' ', '_')), image)
                for image in product_images]
            url_path = request.META['HTTP_HOST']
            image_url = ['http://' + url_path + '/media/' + image for image in product_image]
            items = ['CLOTHE', 'ELECTRONICS', 'BOOKS', 'BEAUTY', 'MEN ACCESSORIES', 'WOMEN ACCESSORIES', 'FURNITURE',
                     'GARDEN']
            sub_items = ['BOY', 'GIRL', 'MEN', 'WOMEN', 'MOBILE', 'LAPTOP', 'TABLET', 'COMPUTER', 'TV', 'CAMERA',
                         'HOME AUDIO', 'FINANCE & ACCOUNTING EXAMS BOOKS', 'GOVERNMENT EXAMS BOOKS',
                         'EXAMS BY UPSC BOOKS', 'ENGINEERING ENTRANCE BOOKS', 'DEFENCE BOOKS',
                         'BANKING & INSURANCE BOOKS', 'ARTS, DESIGN AND EDUCATION BOOKS', 'MAKEUP', 'SKIN', 'HAIR',
                         'FRAGRANCES', 'MENS GROOMING', 'UNISEX PERSONAL CARE', 'BELT', 'CAPS & HATS',
                         'MUFFLERS, SCARVES & GLOVES', 'HANDKERCHIEFS', 'SOCKS', 'SUNGLASSES', 'WALLET', 'WATCHES',
                         'HELMET', 'JEWELLERY', 'BELTS', 'CAPS & HATS', 'HAIR ACCESSORIES', 'SOCKS', 'SUNGLASSES',
                         'UMBRELLAS', 'WATCHES', 'PINS', 'FURNITURE', 'WATERING CANS', 'GARDENING TOOLS',
                         'HOSE NOZZLES', 'GARDEN FAUCETS', 'WATER PUMPS', 'GARDEN SPRAY']
            if product_category not in items:
                return JsonResponse({'message': 'Category Is not Valid'})
            if product_sub_category not in sub_items:
                return JsonResponse({'message': 'Sub Category Is not valid'})
            product_add = Product(
                product_category=product_category,
                product_sub_category=product_sub_category,
                product_images=image_url,
                product_SKU=product_SKU,
                product_name=product_name,
                product_price=product_price,
                product_sale_price=product_sale_price,
                product_quantity=product_quantity,
                product_branding=product_branding,
                product_tags=product_tags,
                product_size=product_size,
                product_color=product_color,
                product_fabric=product_fabric,
                product_description=product_description,
                product_date=datetime.now(),
                product=product,
                product_seller=seller_user,
            )
            product_add.save()
            return JsonResponse({'message': 'Your Product Added successfully.'}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__()}, status=400)


# ==========This function is to insert the bulk product==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def bulk_upload_catalog(request):
    product_category = request.POST['product_category'].upper()
    product_sub_category = request.POST['product_sub_category'].upper()
    product_image = request.FILES.getlist('product_image')
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            file_name = random.randint(00000, 99999)
            workbook = xlsxwriter.Workbook(f'D:\\Task\\E-Com\\media\\file\\Bulk-Catalog-{file_name}.xlsx')
            worksheet = workbook.add_worksheet('Catalog Data')
            worksheet.set_column_pixels(0, 16, 300)
            worksheet.set_default_row(20)
            product_image = [
                default_storage.save(os.path.join(settings.MEDIA_ROOT, 'Product', image.name.replace(' ', '_')), image)
                for image in product_image
            ]
            url_path = request.META['HTTP_HOST']
            image_url = ['http://' + url_path + '/media/' + image for image in product_image]
            expenses = (
                'Images1', 'Images2', 'Images3', 'Images4', 'SKU', 'Name', 'Price', 'Sale Price', 'Quantity',
                'product_category', 'product_sub_category', 'Branding',
                'Tags', 'Size', 'Color', 'Fabric', 'Description')
            row = 1
            for col_num, value in enumerate(expenses):
                for index, image in enumerate(image_url):
                    if value == 'product_category':
                        worksheet.write(0, col_num, value)
                        worksheet.write(row + index, col_num, product_category)
                    elif value == 'product_sub_category':
                        worksheet.write(0, col_num, value)
                        worksheet.write(row + index, col_num, product_sub_category)
                    elif value == 'Images1':
                        worksheet.write(0, col_num, value)
                        worksheet.write(row + index, col_num, image)
                    worksheet.write(0, col_num, value)
            url_path = request.META['HTTP_HOST']
            file_url = ['http://' + url_path + '/media/file/' + f'Bulk-Catalog-{file_name}.xlsx']
            workbook.close()
            return JsonResponse(
                {'message': 'This is a file for uploading catalogs in bulk', 'execl_file_link': file_url},
                status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'execl_file_link': None}, status=400)


# ==========This function is for taking the link of the image==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def get_image_link(request):
    upload_image = request.FILES.getlist('upload_image')
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            product_image = [
                default_storage.save(os.path.join(settings.MEDIA_ROOT, 'Product', image.name.replace(' ', '_')), image)
                for image in upload_image]
            url_path = request.META['HTTP_HOST']
            image_link = ['http://' + url_path + '/media/' + image for index, image in enumerate(product_image)]
            return JsonResponse({'message': 'This is your image link', 'image_link': image_link}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'image_link': None}, status=400)


# ==========This function is for uploading bulk product files==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def upload_catalog_file(request):
    upload_file = request.FILES['upload_file']
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            directory = 'D:\\Task\\E-Com\\media\\file'
            file_path = os.path.join(directory, str(upload_file))
            if os.path.exists(file_path):
                workbook = openpyxl.load_workbook(upload_file)
                worksheet = workbook.active
                max_row = worksheet.max_row
                for row in range(2, max_row + 1):
                    row = worksheet[row]
                    image_links = []
                    for cell in row:
                        if cell.value and ('http' in str(cell.value) in str(cell.value)):
                            image_links.append(str(cell.value))
                    product_id = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "A", "B", "C", "D", "E", "F", "G",
                                  "H"]
                    products_id = random.choices(product_id, k=9)
                    product = "".join(products_id)
                    product_add = Product()
                    product_add.product_images = image_links
                    product_add.product_SKU = row[4].value
                    product_add.product_name = row[5].value
                    product_add.product_price = row[6].value
                    product_add.product_sale_price = row[7].value
                    product_add.product_quantity = row[8].value
                    product_add.product_category = row[9].value
                    product_add.product_sub_category = row[10].value
                    product_add.product_branding = row[11].value
                    product_add.product_tags = row[12].value
                    product_add.product_size = row[13].value
                    product_add.product_color = row[14].value
                    product_add.product_fabric = row[15].value
                    product_add.product_description = row[16].value
                    product_add.product_date = datetime.now()
                    product_add.product = product
                    product_add.product_seller = seller_user
                    product_add.save()
                return JsonResponse({'message': 'Your Bulk Catalog file uploaded Successfully.'}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__()}, status=400)


# ==========This function is for viewing all products==========

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def view_all_product(request):
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            view_product = Product.objects.filter(product_seller=seller_user)
            serializer = ProductSerializer(view_product, many=True)
            for data in serializer.data:
                if data.get('product_quantity') == 0:
                    data['product_stock'] = 'Out of Stock'
                else:
                    data['product_stock'] = 'Is Stock'
                product = Product.objects.get(product_key=data.get('product_key'))
                product.product_stock = data['product_stock']
                product.save()
                data.pop('product_key')
                data.pop('product_sale_price')
                data.pop('product_quantity')
                data.pop('product_sub_category')
                data.pop('product_branding')
                data.pop('product_tags')
                data.pop('product_fabric')
                data.pop('product_description')
                data.pop('product')
                data.pop('product_seller')
            return JsonResponse({'message': 'View all Product', 'product_data': serializer.data}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'product_data': None}, status=400)


# ==========This function is to update the product==========

def product_update(request, update_inventory_catalog, product_images, product_SKU, product_name, product_branding,
                   product_tags,
                   product_color, product_fabric, product_description):
    if len(product_images) != 0:
        product_image = [
            default_storage.save(os.path.join(settings.MEDIA_ROOT, 'Product', image.name.replace(' ', '_')), image)
            for image in product_images]
        url_path = request.META['HTTP_HOST']
        image_url = ['http://' + url_path + '/media/' + image for image in product_image]
    else:
        image_url = update_inventory_catalog.product_images
    if not product_images == '':
        update_inventory_catalog.product_images = image_url
    if not product_SKU == '':
        update_inventory_catalog.product_SKU = product_SKU
    if not product_name == '':
        update_inventory_catalog.product_name = product_name
    if not product_branding == '':
        update_inventory_catalog.product_branding = product_branding
    if not product_tags == '':
        update_inventory_catalog.product_tags = product_tags
    if not product_color == '':
        update_inventory_catalog.product_color = product_color
    if not product_fabric == '':
        update_inventory_catalog.product_fabric = product_fabric
    if not product_description == '':
        update_inventory_catalog.product_description = product_description
    update_inventory_catalog.save()
    return update_inventory_catalog


# ==========This function is to take data to update the product==========

@api_view(['PUT'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def update_product(request):
    product_key = request.POST['product_key']
    product_images = request.FILES.getlist('product_images', '')
    product_SKU = request.POST.get('product_SKU', '')
    product_name = request.POST.get('product_name', '')
    product_branding = request.POST.get('product_branding', '')
    product_tags = request.POST.get('product_tags', '')
    product_color = request.POST.get('product_color', '')
    product_fabric = request.POST.get('product_fabric', '')
    product_description = request.POST.get('product_description', '')
    try:
        user = request.user
        if user:
            symbol = "~", "!", "#", "$", "%", "^", "&", "*", "@", "-"
            for char in symbol:
                if product_key.find(char) != -1:
                    return JsonResponse({'message': 'ID is not valid'}, status=400)
            if not product_key.isdigit():
                return JsonResponse({'message': 'ID is not valid'}, status=400)
            seller_user = Register.objects.get(register_user=user)
            update_inventory_catalog = Product.objects.get(product_seller=seller_user, product_key=product_key)
            product_update_data = {
                'product_data': product_update(request, update_inventory_catalog, product_images, product_SKU,
                                               product_name, product_branding, product_tags, product_color,
                                               product_fabric, product_description)}
            return JsonResponse({'message': 'Your Product Updated Successfully.'}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Product.DoesNotExist:
        return JsonResponse({'message': 'ID is not exist'}, status=400)
    except Exception as e:
        return JsonResponse({'message': e.__str__()}, status=400)


# ==========This function is for deleting the product==========

@api_view(['DELETE'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def delete_product(request):
    product_key = request.POST['product_key']
    try:
        user = request.user
        if user:
            symbol = "~", "!", "#", "$", "%", "^", "&", "*", "@", "-"
            for char in symbol:
                if product_key.find(char) != -1:
                    return JsonResponse({'message': 'ID is not valid'}, status=400)
            if not product_key.isdigit():
                return JsonResponse({'message': 'ID is not valid'}, status=400)
            seller_user = Register.objects.get(register_user=user)
            product_delete = Product.objects.get(product_seller=seller_user, product_key=product_key)
            product_delete.delete()
            return JsonResponse({'message': 'Delete Product Successfully.'}, status=200)
        else:
            return JsonResponse({'message': 'Login First'}, status=401)
    except Product.DoesNotExist:
        return JsonResponse({'message': 'ID is not exist'}, status=400)
    except Exception as e:
        return JsonResponse({'message': e.__str__()}, status=400)


# ==========This function is for inventory==========

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def inventory(request):
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            view_catalog = Product.objects.filter(product_seller=seller_user)
            serial = ProductSerializer(view_catalog, many=True)
            for data in serial.data:
                if data.get('product_quantity') == 0:
                    data['product_stock'] = 'Out of Stock'
                else:
                    data['product_stock'] = 'Is Stock'
                product = Product.objects.get(product_key=data.get('product_key'))
                product.product_stock = data['product_stock']
                product.save()
                data.pop('product_key')
                data.pop('product_sale_price')
                data.pop('product_category')
                data.pop('product_sub_category')
                data.pop('product_branding')
                data.pop('product_tags')
                data.pop('product_color')
                data.pop('product_fabric')
                data.pop('product_description')
                data.pop('product_date')
                data.pop('product')
                data.pop('product_seller')
            return JsonResponse({'message': 'Filter Category', 'inventory_stock_data': serial.data}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'inventory_stock_data': None}, status=400)


# ==========This function is to edit the stock of inventory==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def edit_stock(request):
    inventory_key = request.POST['inventory_key']
    product_stock = request.POST['product_stock']
    try:
        user = request.user
        if user:
            symbol = "~", "!", "#", "$", "%", "^", "&", "*", "@", "-"
            for char in symbol:
                if inventory_key.find(char) != -1:
                    return JsonResponse({'message': 'ID is not valid'}, status=400)
            if not inventory_key.isdigit():
                return JsonResponse({'message': 'ID is not valid'}, status=400)
            seller_user = Register.objects.get(register_user=user)
            stock_edit = Product.objects.get(product_seller=seller_user, product_key=inventory_key)
            stock_edit.product_quantity = product_stock
            stock_edit.save()
            return JsonResponse({'message': 'Stock Edit Successfully.'}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Product.DoesNotExist:
        return JsonResponse({'message': 'ID is not exist.'}, status=400)
    except Exception as e:
        return JsonResponse({'message': e.__str__()}, status=400)


# ==========This function is for filtering inventory data by category==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def inventory_filter_category(request):
    category = request.POST['category'].upper()
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            view_category_filter = Product.objects.filter(product_seller=seller_user, product_category=category)
            serial = ProductSerializer(view_category_filter, many=True)
            if len(view_category_filter) != 0:
                for data in serial.data:
                    data.pop('product_key')
                    data.pop('product_sale_price')
                    data.pop('product_category')
                    data.pop('product_sub_category')
                    data.pop('product_branding')
                    data.pop('product_tags')
                    data.pop('product_color')
                    data.pop('product_fabric')
                    data.pop('product_description')
                    data.pop('product_date')
                    data.pop('product')
                    data.pop('product_seller')
                return JsonResponse({'message': 'Filter Category', 'inventory_filter_category_data': serial.data},
                                    status=200)
            else:
                return JsonResponse({'message': 'Category is Not exist', 'inventory_filter_category_data': None},
                                    status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'inventory_filter_category_data': None}, status=400)


# ==========This function is for editing product inventory==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def inventory_edit_catalog(request):
    inventory_key = request.POST['inventory_key']
    product_images = request.FILES.getlist('product_images', '')
    product_SKU = request.POST.get('product_SKU', '')
    product_name = request.POST.get('product_name', '')
    product_branding = request.POST.get('product_branding', '')
    product_tags = request.POST.get('product_tags', '')
    product_color = request.POST.get('product_color', '')
    product_fabric = request.POST.get('product_fabric', '')
    product_description = request.POST.get('product_description', '')
    try:
        user = request.user
        if user:
            symbol = "~", "!", "#", "$", "%", "^", "&", "*", "@", "-"
            for char in symbol:
                if inventory_key.find(char) != -1:
                    return JsonResponse({'message': 'ID is not valid'}, status=400)
            if not inventory_key.isdigit():
                return JsonResponse({'message': 'ID is not valid'}, status=400)
            seller_user = Register.objects.get(register_user=user)
            update_inventory_catalog = Product.objects.get(product_seller=seller_user, product_key=inventory_key)
            inventory_update_data = {
                'product_data': product_update(request, update_inventory_catalog, product_images, product_SKU,
                                               product_name, product_branding, product_tags, product_color,
                                               product_fabric, product_description)}
            return JsonResponse({'message': 'Edit Catalog Successfully.'}, status=200)
        else:
            return JsonResponse({'message': 'Login First'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__()}, status=400)


# ==========This function is for product rating==========

def rating_product(product_rating, products):
    rating_dict = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    rating_data = []
    for rating in product_rating:
        rating_dict[rating.feedback_rating] += 1
    total = sum(rating_dict.values())
    if total != 0:
        rating_average = sum(key * value for key, value in rating_dict.items()) / total
    else:
        rating_average = 0
    data = {
        'images': products.product_images,
        'name': products.product_name,
        'product_id': products.product,
        'category': products.product_category,
        'product_rating': rating_average,
        'total_rating': total,
        'very_bad': rating_dict[1],
        'bad': rating_dict[2],
        'ok-ok': rating_dict[3],
        'good': rating_dict[4],
        'very_good': rating_dict[5]
    }
    rating_data.append(data)
    return rating_data


# ==========This function is for viewing the product's rating==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def view_rating(request):
    inventory_key = request.POST['inventory_key']
    try:
        user = request.user
        if user:
            symbol = "~", "!", "#", "$", "%", "^", "&", "*", "@", "-"
            for char in symbol:
                if inventory_key.find(char) != -1:
                    return JsonResponse({'message': 'ID is not valid'}, status=400)
            if not inventory_key.isdigit():
                return JsonResponse({'message': 'ID is not valid'}, status=400)
            seller_user = Register.objects.get(register_user=user)
            products = Product.objects.get(product_key=inventory_key, product_seller=seller_user)
            product_rating = BuyerFeedback.objects.filter(feedback_product=products)
            rating_data = rating_product(product_rating, products)
            return JsonResponse({'message': 'View Rating', 'rating_data': rating_data}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Product.DoesNotExist:
        return JsonResponse({'message': 'Product is not exist'}, status=400)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'rating_data': None}, status=400)


# ==========This function is for product data==========

def order_data(view_all_order):
    view_order = [{'order_id': order.order,
                   'image': order.payment.details.product.product_images,
                   'name': order.payment.details.product.product_name,
                   'product_sku': order.payment.details.product.product_SKU,
                   'company_id': order.company,
                   'quantity': order.qty,
                   'size': order.product_size,
                   'dispatch_date': order.dispatch_date}
                  for order in view_all_order]
    return view_order


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def pending_order(request):
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            buyer_order = BuyerPayment.objects.filter(cancel=True)
            buyer_all_order = Order.objects.filter(payment__details__product__product_seller=seller_user,
                                                   payment__cancel=True).only('order')
            order_list = [order.order for order in buyer_all_order]
            for order in buyer_order:
                if order.order_key not in order_list:
                    company_id1 = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "A", "B", "C", "D", "E", "F", "G",
                                   "H"]
                    company_id2 = random.choices(company_id1, k=9)
                    company = "".join(company_id2)
                    orders = Order(
                        payment=order,
                        product=order.details.product,
                        buyer=order.buyer,
                        qty=order.details.qty,
                        product_color=order.details.product.product_color,
                        product_size=order.details.product.product_size,
                        total=order.amount,
                        order=order.order_key,
                        company=company,
                        order_date=datetime.now(),
                    )
                    orders.dispatch_date = orders.order_date + timedelta(days=4)
                    orders.save()
            view_all_order = Order.objects.filter(payment__details__product__product_seller=seller_user, status=True)
            view_order_data = order_data(view_all_order)
            return JsonResponse({'message': 'View All Pending Order', 'view_all_order_data': view_order_data},
                                status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'view_all_order_data': None}, status=400)


# ==========This function is for filtering by order data==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def filter_order_date(request):
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')
    start_time = time(00, 00, 00)
    end_time = time(23, 59, 59)
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            start_datetime = datetime.strptime(f'{start_date} {start_time}', "%Y-%m-%d %H:%M:%S")
            end_datetime = datetime.strptime(f'{end_date} {end_time}', "%Y-%m-%d %H:%M:%S")
            view_all_order = Order.objects.filter(order_date__range=(start_datetime, end_datetime),
                                                  payment__details__product__product_seller=seller_user,
                                                  payment__cancel=True, status=True)
            view_filter_data = order_data(view_all_order)
            return JsonResponse({'message': 'Filter By: Order Date', 'filter_order_date': view_filter_data}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except ValueError:
        return JsonResponse({'message': 'Date is not valid'}, status=400)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'filter_order_date': None}, status=400)


# ==========This function is for filtering by dispatch data==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def filter_dispatch_date(request):
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')
    start_time = time(00, 00, 00)
    end_time = time(23, 59, 59)
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            start_datetime = datetime.strptime(f'{start_date} {start_time}', "%Y-%m-%d %H:%M:%S")
            end_datetime = datetime.strptime(f'{end_date} {end_time}', "%Y-%m-%d %H:%M:%S")
            view_all_order = Order.objects.filter(dispatch_date__range=(start_datetime, end_datetime),
                                                  payment__details__product__product_seller=seller_user,
                                                  payment__cancel=True, status=True)
            view_filter_data = order_data(view_all_order)
            return JsonResponse({'message': 'Filter By: Dispatch Date', 'filter_order_date': view_filter_data},
                                status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except ValueError:
        return JsonResponse({'message': 'Date is not valid'}, status=400)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'filter_order_date': None}, status=400)


# ==========This function is for searching orders==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def order_search(request):
    search = request.POST.get('search', '')
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            view_all_order = Order.objects.filter(
                Q(order__icontains=search) | Q(company__icontains=search) | Q(
                    payment__details__product__product_SKU__icontains=search),
                payment__details__product__product_seller=seller_user, status=True)
            if view_all_order.count() != 0 and search != '':
                search_data = order_data(view_all_order)
                return JsonResponse(
                    {'message1': 'Search Order', 'message2': 'Search Order Only SKU, Order ID, Company ID',
                     'search_order_data': search_data}, status=200)
            else:
                return JsonResponse(
                    {'Message': 'Order Is Not Found', 'Message2': 'Search Order Only SKU, Order ID, Company ID',
                     'search_order_data': None}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'search_order_data': None}, status=400)


# ==========This function is for data of order accept==========

def accept_data(accept):
    accept_order = Accept(
        order=accept,
        product=accept.product,
        buyer=accept.buyer,
        qty=accept.qty,
        product_color=accept.product_color,
        product_size=accept.product_size,
        total=accept.total
    )
    accept_order.save()
    accept.status = False
    accept.save()
    product = Product.objects.get(product_key=accept.product.product_key)
    product.product_quantity = product.product_quantity - accept.qty
    product.save()
    return accept


# ==========This function is for accepting orders==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def order_accept(request):
    order_key = request.POST['order_key']
    key_list = order_key.split(',')
    try:
        user = request.user
        if user:
            symbol = "~", "!", "#", "$", "%", "^", "&", "*", "@"
            for char in symbol:
                if order_key.find(char) != -1:
                    return JsonResponse({'message': 'ID is not valid'}, status=400)
            if order_key.isalpha():
                return JsonResponse({'message': 'ID is not valid'}, status=400)
            seller_user = Register.objects.get(register_user=user)
            if not len(key_list) > 1:
                accept = Order.objects.get(order_key=order_key, status=True,
                                           payment__details__product__product_seller=seller_user)
                accept_order = accept_data(accept)
                return JsonResponse({'message': 'Order Accept'}, status=200)
            else:
                for key in key_list:
                    accept = Order.objects.get(order_key=key, status=True,
                                               payment__details__product__product_seller=seller_user)
                    accept_order = accept_data(accept)
                return JsonResponse({'message': 'Order Accept'}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Order.DoesNotExist:
        return JsonResponse({'message': 'ID is not exist.'}, status=400)
    except Exception as e:
        return JsonResponse({'message': e.__str__()}, status=400)


# ==========This function is for viewing accepted orders==========

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def ready_to_ship(request):
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            view_accept_all_order = Accept.objects.filter(status=True, order__product__product_seller=seller_user)
            accept_orders_data = [{
                'order_id': order.order.order,
                'image': order.product.product_images,
                'name': order.product.product_name,
                'product_SKU': order.product.product_SKU,
                'company_id': order.order.company,
                'quantity': order.qty,
                'size': order.product.product_size,
                'dispatch_date': f"{order.order.dispatch_date.day} {order.order.dispatch_date.strftime('%b')}'{order.order.dispatch_date.strftime('%y')}"
            } for order in view_accept_all_order]
            return JsonResponse({'message': 'Ready To Ship', 'accept_order_data': accept_orders_data}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'accept_order_data': None}, status=400)


# ==========This function is for shipping label data==========

def label_data(label, seller_user, purchase_order_no, invoice_no, invoice_date):
    tax, gst1, gst2, gst3 = 0, 5.0, 12.0, 18.0
    if label.order.product.product_category in ['BOOKS', 'BEAUTY', 'FURNITURE', 'GARDEN']:
        tax = label.total * gst1 / 100
    elif label.order.product.product_category in ['CLOTHE', 'MEN ACCESSORIES', 'WOMEN ACCESSORIES']:
        tax = label.total * gst2 / 100
    elif label.order.product.product_category == 'ELECTRONICS':
        tax = label.total * gst3 / 100
    total = tax + label.total
    data = {
        'customer_name': label.buyer.buyer.first_name + ' ' + label.buyer.buyer.last_name,
        'customer_address': label.buyer.user_address,
        'seller_name': seller_user.register_user.first_name + seller_user.register_user.last_name,
        'seller_address': seller_user.address,
        'product_details': {'product_SKU': label.product.product_SKU, 'size': label.product.product_size,
                            'qty': label.qty,
                            'color': label.product_color, 'order_no': label.order.order},
        'tax_invoice': {'ship_to': (
            label.order.payment.details.checkout.ord_rec_name, label.order.payment.details.checkout.street_address,
            label.order.payment.details.checkout.apartment_address, label.order.payment.details.checkout.pincode),
            'sold_by': (
                seller_user.register_user.first_name + " " + seller_user.register_user.last_name, seller_user.address),
            'purchase_order_no': purchase_order_no, 'invoice_no': invoice_no, 'order_date': label.order.order_date,
            'invoice_date': invoice_date},
        'description': label.product.product_name, 'qty': label.qty, 'gross_amount': total,
        'taxable_value': label.total, 'taxes': f'GST RS.{tax}', 'total': total
    }
    return data


# ==========This function is for shipping labels==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def shipping_label(request):
    order_key = request.POST['order_key']
    key_list = order_key.split(',')
    try:
        user = request.user
        if user:
            symbol = "~", "!", "#", "$", "%", "^", "&", "*", "@"
            for char in symbol:
                if order_key.find(char) != -1:
                    return JsonResponse({'message': 'ID is not valid'}, status=400)
            if order_key.isalpha():
                return JsonResponse({'message': 'ID is not valid'}, status=400)
            seller_user = Register.objects.get(register_user=user)
            purchase = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '0']
            invoice = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j']
            purchase_order = random.choices(purchase, k=12)
            purchase_order_no = "".join(purchase_order)
            invoice_key = random.choices(purchase + invoice, k=11)
            invoice_no = "".join(invoice_key)
            invoice_date = date.today()
            if len(key_list) == 1:
                label = Accept.objects.get(accept_key=order_key, status=True,
                                           product__product_seller=seller_user)
                data = label_data(label, seller_user, purchase_order_no, invoice_no, invoice_date)
                return JsonResponse({'message': 'Shipping Label', 'label_data': data}, status=200)
            else:
                data_list = []
                for key in key_list:
                    label = Accept.objects.get(accept_key=key, status=True, product__product_seller=seller_user)
                    data = label_data(label, seller_user, purchase_order_no, invoice_no, invoice_date)
                    data_list.append(data)
                return JsonResponse({'message': 'Shipping Label', 'label_data': data_list}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Accept.DoesNotExist:
        return JsonResponse({'message': 'ID is not exist'}, status=400)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'label_data': None}, status=400)


# ==========This function is for data of order cancel==========

def cancel_data(cancel):
    order_cancel = Cancel(
        order=cancel,
        product=cancel.product,
        buyer=cancel.buyer,
        qty=cancel.qty,
        product_color=cancel.product_color,
        product_size=cancel.product_size,
        total=cancel.total
    )
    order_cancel.save()
    cancel.status = False
    cancel.save()
    return cancel


# ==========This function is for canceling orders==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cancel_order(request):
    order_key = request.POST['order_key']
    key_list = order_key.split(',')
    try:
        user = request.user
        if user:
            symbol = "~", "!", "#", "$", "%", "^", "&", "*", "@", " "
            for char in symbol:
                if order_key.find(char) != -1:
                    return JsonResponse({'message': 'ID is not valid'}, status=400)
            if order_key.isalpha():
                return JsonResponse({'message': 'ID is not valid'}, status=400)
            seller_user = Register.objects.get(register_user=user)
            if not len(key_list) > 1:
                cancel = Order.objects.get(order_key=order_key, status=True,
                                           payment__details__product__product_seller=seller_user)
                order_cancel = cancel_data(cancel)
                return JsonResponse({'message': 'Order Cancel'}, status=200)
            else:
                for key in key_list:
                    cancel = Order.objects.get(order_key=key, status=True,
                                               payment__details__product__product_seller=seller_user)
                    order_cancel = cancel_data(cancel)
                return JsonResponse({'message': 'Order Cancel'}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Order.DoesNotExist:
        return JsonResponse({'message': 'ID is not exist.'}, status=400)
    except Exception as e:
        return JsonResponse({'message': e.__str__()}, status=400)


# ==========This function is for viewing cancelled orders==========

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def cancelled(request):
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            view_cancel_all_order = Cancel.objects.filter(status=True, order__product__product_seller=seller_user)
            cancel_orders_data = [{
                'image': order.product.product_images,
                'name': order.product.product_name,
                'order_id': order.order.order,
                'product_SKU': order.product.product_SKU,
                'company_id': order.order.company,
                'quantity': order.qty,
                'size': order.product.product_size,
            } for order in view_cancel_all_order]
            return JsonResponse({'message': 'View Cancelled', 'cancel_order_data': cancel_orders_data}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'cancel_order_data': None}, status=400)


# ==========This function is for pricing data==========

def pricing_data(product, seller_user):
    view_product = []
    for data in product:
        orders = Order.objects.filter(product=data, product__product_seller=seller_user)
        total, total_order = 0, 0
        order_comprehension = [
            (total := total + j.total, total_order := total_order + j.qty)
            for j in orders
        ]
        data = {
            'product_details': {'image': data.product_images, 'name': data.product_name,
                                'product_SKU': data.product_SKU,
                                'size': data.product_size},
            'current_stock': data.product_quantity,
            'growth': {'orders': total_order, 'sales': f'₹{total}'},
            'current_customer_price': data.product_price
        }
        view_product.append(data)
    return view_product


# ==========This function is to view the price of the product==========

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def pricing(request):
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            product = Product.objects.filter(product_seller=seller_user)
            view_product_data = pricing_data(product, seller_user)
            return JsonResponse({'message': 'Pricing', 'product_data': view_product_data}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'product_data': None}, status=400)


# ==========This function is for changing the price of the product==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def edit_pricing(request):
    pricing_key = request.POST['pricing_key']
    product_price = request.POST.get('product_price')
    product_sale_price = request.POST.get('product_sale_price')
    try:
        user = request.user
        if user:
            symbol = "~", "!", "#", "$", "%", "^", "&", "*", "@", " ", ","
            for char in symbol:
                if pricing_key.find(char) != -1 or product_price.find(char) != -1 or product_sale_price.find(
                        char) != -1:
                    return JsonResponse({'message': 'ID is not valid'}, status=400)
            if pricing_key.isalpha():
                return JsonResponse({'message': 'ID is not valid'}, status=400)
            seller_user = Register.objects.get(register_user=user)
            edit_product_price = Product.objects.get(product_key=pricing_key, product_seller=seller_user)
            if not product_price == '':
                edit_product_price.product_price = product_price
            if not product_sale_price == '':
                edit_product_price.product_sale_price = product_sale_price
            edit_product_price.save()
            return JsonResponse({'message': 'Price Edit Successfully'}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Product.DoesNotExist:
        return JsonResponse({'message': 'ID is not exist.'}, status=400)
    except Exception as e:
        return JsonResponse({'message': e.__str__()}, status=400)


# ==========This function is for filtering by product category==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def filter_category(request):
    category = request.POST['category'].upper()
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            filter_product = Product.objects.filter(product_seller=seller_user, product_sub_category=category)
            symbol = "~", "!", "#", "$", "%", "^", "&", "*", "@"
            for char in symbol:
                if category.find(char) != -1:
                    return JsonResponse({'message': 'Category is not valid'}, status=400)
            if category.isdigit() or len(filter_product) == 0:
                return JsonResponse({'message': 'Category is not valid'}, status=400)
            view_product_data = pricing_data(filter_product, seller_user)
            return JsonResponse({'message': 'Filter By: Category', 'product_data': view_product_data}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'Product data': None}, status=400)


# ==========This function is for filtering by product data==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def date_growth(request):
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')
    start_time = time(00, 00, 00)
    end_time = time(23, 59, 59)
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            start_datetime = datetime.strptime(f'{start_date} {start_time}', "%Y-%m-%d %H:%M:%S")
            end_datetime = datetime.strptime(f'{end_date} {end_time}', "%Y-%m-%d %H:%M:%S")
            filter_product = Product.objects.filter(product_date__range=(start_datetime, end_datetime),
                                                    product_seller=seller_user)
            view_product_data = pricing_data(filter_product, seller_user)
            return JsonResponse({'message': 'Date Growth', 'growth_product_data': view_product_data}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except ValueError:
        return JsonResponse({'message': 'Date is not valid'}, status=400)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'growth_product_data': None}, status=400)


# ==========This function is for data of order return==========

def return_data(return_order):
    view_returns = [{
        'images': order.order.details.product.product_images,
        'name': order.order.details.product.product_name,
        'product_SKU': order.order.details.product.product_SKU,
        'category': order.order.details.product.product_category,
        'qty': order.order.details.qty,
        'size': order.order.details.product.product_size,
        'order_id': order.order.order_key,
        'return _reason': order.order_return_message,
        'return_shipping_fee': f'₹{order.return_shipping_Fee}',
        'created_date': f"{order.return_date.day} {order.return_date.strftime('%b')}'{order.return_date.strftime('%y')}",
        'return_id': order.returns,
    } for order in return_order]
    return view_returns


# ==========This function is for order return tracking==========

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def return_tracking(request):
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            return_order = Return.objects.filter(order__details__product__product_seller=seller_user)
            view_returns_data = return_data(return_order)
            return JsonResponse({'message': 'Tracking Return', 'return_product_data': view_returns_data}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'return_product_data': None}, status=400)


# ==========This function is for overview of order returns==========

@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def return_overview(request):
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            product = Product.objects.filter(product_seller=seller_user)
            view_returns_data = []
            for data in product:
                return_order = Return.objects.filter(order__details__product__product_seller=seller_user,
                                                     order__details__product__product=data.product)
                order = Accept.objects.filter(product__product=data.product, product__product_seller=seller_user)
                total, total_order = 0, 0
                for k in order:
                    total_order += k.qty
                for _ in return_order:
                    total += 1
                if total != 0 and total_order != 0:
                    return_order_total = total
                    return_returns_total = round(total * 100 / total_order, 1)
                else:
                    return_order_total = '-'
                    return_returns_total = '0.0'
                data = {
                    'image': data.product_images,
                    'name': data.product_name,
                    'category': data.product_category,
                    'return_order': return_order_total,
                    'returns': f'{return_returns_total}%',
                }
                view_returns_data.append(data)

            return JsonResponse({'message': 'Return Over View', 'returns_data': view_returns_data}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'returns_data': None}, status=400)


# ==========This function returns order to filter by category==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def return_filter_category(request):
    category = request.POST['category'].upper()
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            return_order = Return.objects.filter(order__details__product__product_seller=seller_user,
                                                 order__details__product__product_category=category)
            symbol = "~", "!", "#", "$", "%", "^", "&", "*", "@"
            for char in symbol:
                if category.find(char) != -1:
                    return JsonResponse({'message': 'Category is not valid'}, status=400)
            if category.isdigit():
                return JsonResponse({'message': 'Category is not valid'}, status=400)
            view_returns_data = return_data(return_order)
            return JsonResponse({'message': 'Filter By: Category', 'filter_data': view_returns_data}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'filter_data': None}, status=400)


# ==========This function is for filtering the return order from the return file==========

@api_view(['POST'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def return_filter_date(request):
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')
    start_time = time(00, 00, 00)
    end_time = time(23, 59, 59)
    try:
        user = request.user
        if user:
            seller_user = Register.objects.get(register_user=user)
            start_datetime = datetime.strptime(f'{start_date} {start_time}', "%Y-%m-%d %H:%M:%S")
            end_datetime = datetime.strptime(f'{end_date} {end_time}', "%Y-%m-%d %H:%M:%S")
            return_order = Return.objects.filter(return_date__range=(start_datetime, end_datetime),
                                                 order__details__product__product_seller=seller_user)
            view_returns_data = return_data(return_order)
            return JsonResponse({'message': 'Filter By: Return Date', 'filter_data': view_returns_data}, status=200)
        else:
            return JsonResponse({'message': 'If you are not currently logged in, please login first'}, status=401)
    except ValueError:
        return JsonResponse({'message': 'Date is not valid'}, status=400)
    except Exception as e:
        return JsonResponse({'message': e.__str__(), 'filter_data': None}, status=400)
